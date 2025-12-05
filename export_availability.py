"""
Exporta disponibilidad de Museos Vaticanos a Excel
"""
import sys
import io
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from vatican_client import VaticanClient
from config import DEFAULT_VISIT_TAG, DEFAULT_VISITOR_NUM, DEFAULT_WHO_ID, PRODUCT_FILTER


def export_to_excel(output_file: str = None, max_days: int = 60):
    """
    Consulta disponibilidad de las fechas y exporta a Excel.

    Args:
        output_file: Nombre del archivo Excel de salida
        max_days: Máximo de días a consultar (por defecto 60, para evitar errores con fechas lejanas)
    """
    if not output_file:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'disponibilidad_vaticano_{timestamp}.xlsx'

    print("Iniciando cliente Vatican...")
    client = VaticanClient()

    print("Obteniendo fechas disponibles...")
    all_dates = client.get_available_dates(DEFAULT_VISIT_TAG)

    # Limitar a los próximos max_days días para evitar errores 500 con fechas lejanas
    today = datetime.now().date()
    dates = []
    for date_str in all_dates:
        try:
            date_parts = date_str.split('/')
            date_obj = datetime(int(date_parts[2]), int(date_parts[1]), int(date_parts[0])).date()
            if (date_obj - today).days <= max_days:
                dates.append(date_str)
        except:
            continue

    print(f"Fechas abiertas encontradas: {len(all_dates)}")
    print(f"Consultando las próximas {len(dates)} fechas (hasta {max_days} días)")

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Disponibilidad"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    available_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    low_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    sold_out_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ["Fecha", "Dia", "Producto", "Estado", "Disponible"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    row = 2
    total_available = 0

    for i, date_str in enumerate(dates):
        print(f"Consultando {date_str} ({i+1}/{len(dates)})...")

        # Parsear fecha para obtener dia de la semana
        try:
            date_parts = date_str.split('/')
            date_obj = datetime(int(date_parts[2]), int(date_parts[1]), int(date_parts[0]))
            day_names = ['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes', 'Sabado', 'Domingo']
            day_name = day_names[date_obj.weekday()]
        except:
            day_name = ""

        # Obtener productos disponibles
        products = client.get_available_products(
            date_str,
            visitor_num=DEFAULT_VISITOR_NUM,
            tag=DEFAULT_VISIT_TAG,
            who_id=DEFAULT_WHO_ID,
            product_filter=PRODUCT_FILTER
        )

        if products:
            for product in products:
                name = product.get('name', 'N/A')
                availability = product.get('availability', 'UNKNOWN')
                is_available = availability in ['AVAILABLE', 'LOW_AVAILABILITY']

                ws.cell(row=row, column=1, value=date_str).border = border
                ws.cell(row=row, column=2, value=day_name).border = border
                ws.cell(row=row, column=3, value=name).border = border

                status_cell = ws.cell(row=row, column=4, value=availability)
                status_cell.border = border
                if availability == 'AVAILABLE':
                    status_cell.fill = available_fill
                elif availability == 'LOW_AVAILABILITY':
                    status_cell.fill = low_fill

                avail_cell = ws.cell(row=row, column=5, value="SI" if is_available else "NO")
                avail_cell.border = border
                avail_cell.alignment = Alignment(horizontal='center')
                if is_available:
                    avail_cell.fill = available_fill
                    total_available += 1

                row += 1
        else:
            # No hay productos disponibles para esta fecha
            ws.cell(row=row, column=1, value=date_str).border = border
            ws.cell(row=row, column=2, value=day_name).border = border
            ws.cell(row=row, column=3, value="Sin disponibilidad").border = border

            status_cell = ws.cell(row=row, column=4, value="SOLD_OUT")
            status_cell.border = border
            status_cell.fill = sold_out_fill

            avail_cell = ws.cell(row=row, column=5, value="NO")
            avail_cell.border = border
            avail_cell.alignment = Alignment(horizontal='center')
            avail_cell.fill = sold_out_fill

            row += 1

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12

    # Agregar resumen
    row += 2
    ws.cell(row=row, column=1, value="RESUMEN").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value=f"Total fechas consultadas: {len(dates)}")
    row += 1
    ws.cell(row=row, column=1, value=f"Productos con disponibilidad: {total_available}")
    row += 1
    ws.cell(row=row, column=1, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

    # Guardar
    wb.save(output_file)
    print(f"\nExcel guardado: {output_file}")
    print(f"Total productos disponibles: {total_available}")

    return output_file


if __name__ == '__main__':
    export_to_excel()
